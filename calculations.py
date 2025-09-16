#!/usr/bin/env python3
#this script successfully opens each input xlsx file, iterate over each sheet, add necessary columns, calculate required distance, diff and worst case by date 21.08.25
"""
Bulk Excel processor for clearance‐check sheets.

What it does
------------
1.  Scans an input directory for files that match
      Cond_10C_TR####a###_###.xlsx
    and stores the “start span” ( e.g. TR1730a001 )
    and the “end span” ( e.g. TR1730a002 ) for later use.

2.  For every workbook and every sheet whose name follows
      d1-d2_d3-d4               (e.g. 41-42_43-44)
    it
      • adds the columns  sag, K, LK, C1, C2,
        required-distance, diff, worst-case
      • fills the columns row-by-row exactly as required
        (formulas reproduced below in code comments)
      • writes a processed copy next to the source file
        ( …_processed.xlsx ).

The script is purely local-IO; nothing is sent anywhere.
"""

import argparse
import glob
import os
import re
from typing import Tuple, Dict, Set
import numpy as np

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# ----------------------------------------------------------------------
# helpers
# ----------------------------------------------------------------------
FILE_RE = re.compile(r"Cond_10C_(TR\d{4}a\d{3})_(\d{3})\.xlsx$", re.I)
SHEET_RE = re.compile(r"(\d+)-(\d+)_(\d+)-(\d+)$")

# Circuit membership tables (expand if your network changes)
CIRCUIT_1: Set[int] = {41, 42, 43}
CIRCUIT_2: Set[int] = {44, 45, 46}


EARTH_WIRE = 0          # symbolic circuit ID for EW towers

# ------------------------------------------------------------------
# helper: locate a distance / sag column no matter how it was named
# ------------------------------------------------------------------
def _pick_distance_col(df: pd.DataFrame, a: int, b: int) -> str:
    """
    Return the first column in *df* that matches one of the accepted
    names for the span a-b.  Raises KeyError if none exist.
    Accepted names:
      • "Straight Distance of a-b"
      • "Sag of a-b"
    """
    candidates = [
        f"Straight Distance of {a}-{b}",
        f"Sag of {a}-{b}",
    ]
    for col in candidates:
        if col in df.columns:
            return col
    raise KeyError(f"No distance/sag column found for span {a}-{b}. "
                   f"Tried: {', '.join(candidates)}")



def _which_circuit(tower: int) -> int:
    """
    Return
      1  if tower in CIRCUIT_1
      2  if tower in CIRCUIT_2
      0  otherwise  (earth-wire / unknown circuit)
    """
    if tower in CIRCUIT_1:
        return 1
    if tower in CIRCUIT_2:
        return 2
    return EARTH_WIRE


def _parse_sheet_name(sheet_name: str) -> Tuple[int, int, int, int]:
    """
    Extract digit1, digit2, digit3, digit4 from a name like 41-42_43-44.
    """
    m = SHEET_RE.match(sheet_name)
    if not m:
        raise ValueError(f"Sheet name {sheet_name!r} does not match d1-d2_d3-d4.")
    d1, d2, d3, d4 = map(int, m.groups())
    return d1, d2, d3, d4


def _load_metadata(xl_path: str) -> Tuple[str, str]:
    """
    Extract start & end span codes from the workbook filename.
    Returns (start_span, end_span_suffix).
    """
    m = FILE_RE.search(os.path.basename(xl_path))
    if not m:
        raise ValueError(f"Filename {xl_path!r} does not match expected pattern.")
    return m.group(1), m.group(2)


def process_sheet(df: pd.DataFrame,
                  d1: int, d2: int, d3: int, d4: int) -> pd.DataFrame:
    """
    Add/compute the required columns for a single sheet and return the new DF.
    """
    # ------------------------------------------------------------------
    # 2-3  sag  (max of two distance/sag columns)
    # ------------------------------------------------------------------
    col_sd12 = _pick_distance_col(df, d1, d2)
    col_sd34 = _pick_distance_col(df, d3, d4)

    df["sag"] = df[[col_sd12, col_sd34]].max(axis=1)

    # ------------------------------------------------------------------
    # 2-4  K  =  -(0.1/90)*Beta + 0.75
    # ------------------------------------------------------------------
    df["K"] = -(0.1 / 90.0) * df["Beta Angle (°)"] + 0.75

    # ------------------------------------------------------------------
    # 2-5  constant LK
    # ------------------------------------------------------------------
    df["LK"] = 2.0  # metres

    # ------------------------------------------------------------------
    # 2-6  C1 / C2  by circuit comparison of d1 & d3
    # ------------------------------------------------------------------
    cir1 = _which_circuit(d1)
    cir3 = _which_circuit(d3)

    # Earth-wire rule: if either tower is EW (0) we use C1 = 1.93 m only
    if EARTH_WIRE in (cir1, cir3):
        df["C1"] = 1.93
        df["C2"] = 0.0
        # required distance = C1 only (rule supplied by user)
        df["required-distance"] = df["C1"]
    else:
        # both towers belong to real circuits -> old logic
        same_circuit = cir1 == cir3
        if same_circuit:
            df["C1"] = 1.93
            df["C2"] = 0.0
        else:
            df["C1"] = 0.0
            df["C2"] = 2.29
        df["required-distance"] = (
        df["K"] * np.sqrt(df["sag"] + df["LK"])   # k * sqrt(f + LK)
        + df["C1"] + df["C2"]                     # + C1 or C2
        )

    # ------------------------------------------------------------------
    # 2-8  diff = actual - required
    # ------------------------------------------------------------------
    df["diff"] = df["Distance Between Powerlines"] - df["required-distance"]

    # ------------------------------------------------------------------
    # 2-9  worst-case  (single value, other cells left blank)
    # ------------------------------------------------------------------
    worst_idx = df["diff"].idxmin()      # row where diff is the lowest
    worst_val = df.loc[worst_idx, "diff"]

    df["worst-case"] = np.nan            # start with empty cells
    df.loc[worst_idx, "worst-case"] = worst_val

    return df


def process_workbook(xl_path: str) -> None:
    """
    Open a workbook, process every matching sheet and write to new file.
    """
    start_span, end_suffix = _load_metadata(xl_path)
    print(f"  Processing {os.path.basename(xl_path)}  "
          f"[start span {start_span}, end suffix {end_suffix}]")

    wb = load_workbook(xl_path)
    for sheet_name in wb.sheetnames:
        try:
            d1, d2, d3, d4 = _parse_sheet_name(sheet_name)
        except ValueError:
            # Skip non-conforming sheets silently
            continue

        print(f"    • sheet {sheet_name}")
        ws = wb[sheet_name]
        df = pd.DataFrame(ws.values)
        df.columns = df.iloc[0]        # first row -> header
        df = df[1:].reset_index(drop=True)

        df = process_sheet(df, d1, d2, d3, d4)

        # replace sheet contents with the processed DF
        ws.delete_rows(1, ws.max_row)
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

    # save
    out_path = os.path.splitext(xl_path)[0] + "_processed.xlsx"
    wb.save(out_path)
    print(f"    → written {os.path.basename(out_path)}")


# ----------------------------------------------------------------------
# CLI
# ----------------------------------------------------------------------
def main() -> None:
    ap = argparse.ArgumentParser(
        description="Populate sag / clearance data in Cond_10C workbooks.")
    ap.add_argument("input_folder", help="Folder containing the XLSX source files")
    args = ap.parse_args()

    pattern = os.path.join(args.input_folder, "Cond_10C_*.xlsx")
    files = glob.glob(pattern)
    if not files:
        print("No matching files found.")
        return

    for fp in files:
        try:
            process_workbook(fp)
        except Exception as exc:
            print(f"ERROR in {fp}: {exc}")


if __name__ == "__main__":
    main()
