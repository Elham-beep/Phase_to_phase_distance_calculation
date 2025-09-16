#!/usr/bin/env python3
# run the code using: python calculations.py "C:\path\to\PLS_CADD_file"
#this code runs successfully and creates a new excel file with the required columns and calculations by the time 26.08.25

"""
Bulk Excel processor for clearance‐check sheets.

What it does
------------
1.  Scans an input directory for files that match
      Cond_10C_TR####a###_###.xlsx
    and stores the “start span” ( e.g. TR1730a001 )
    and the “end span” ( e.g. TR1730a002 ) for later use.

2.  For every workbook and every sheet whose name follows
      d1-d2_d3-d4               (e.g. 41-42_43-44)
    it
      • adds the columns  sag, K, LK, C1, C2,
        required-distance, diff, worst-case
      • fills the columns row-by-row exactly as required
        (formulas reproduced below in code comments)
      • writes a processed copy next to the source file
        ( …_processed.xlsx ).
        
3.  map and write the results in a summary workbook "Formulas Distances_rev1.xlsx" in the specified path.

The script is purely local-IO; nothing is sent anywhere.
"""

import argparse
import glob
import os
import re
from typing import Tuple, Set
import numpy as np

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows


# ---------------------------------------------------------------
# CONSTANTS
# ---------------------------------------------------------------
FILE_RE = re.compile(r"Cond_10C_(TR\d{4}a\d{3})_(\d{3})\.xlsx$", re.I)
SHEET_RE = re.compile(r"(\d+)-(\d+)_(\d+)-(\d+)$")
# Hard-coded location for the summary workbook
RESULTS_PATH = os.path.join(os.path.dirname(__file__),r"C:\Users\bimax\DC\ACCDocs\Axpo Grid AG\DEMO_AXPO_Leitungen\Project Files\Grid 4.0 - PLS Distances Development\elham-outputs\Formulas Distances_rev1.xlsx")

# Circuit membership tables (expand if your network changes)
CIRCUIT_1: Set[int] = {41, 42, 43}
CIRCUIT_2: Set[int] = {44, 45, 46}
EARTH_WIRE = 0          # symbolic circuit ID for EW towers

#sheet names 
PH_PH_SHEET = "Result_Dist_Ph-Ph (10ºC)"
PH_EW_SHEET = "Result_Dist_Ph-EW"

# column maps – 1-based indices in the template workbooks
COL_PH_PH = {
    "row":        1,  "src_start": 2,  "src_end":   3,  "src_set":   4,
    "src_phase":  5,  "tgt_start": 6,  "tgt_end":   7,  "tgt_set":   8,
    "tgt_phase":  9,  "station":  10,  "sag":      11,  "lk":       12,
    "weather":   13,  "beta":     14,  "w":        15,  "k":        16,
    "c1":        17,  "c2":       18,  "req":      19,  "cur":      20,
    "ok":        21,
}

COL_PH_EW = {
    "row":        1,  "src_start": 2,  "src_end":   3,  "src_set":   4,
    "src_phase":  5,  "tgt_start": 6,  "tgt_end":   7,  "tgt_set":   8,
    "tgt_phase":  9,  "station":  10,  "wind":     11,  # left blank
    "c1":        12,  "c3":       13,  "req":      14,  "cur":      15,
    "ok":        16,
}


# ------------------------------------------------------------------
# helper: locate a distance / sag column no matter how it was named
# ------------------------------------------------------------------
def _pick_distance_col(df: pd.DataFrame, a: int, b: int) -> str:
    """
    Return the first column in *df* that matches one of the accepted
    names for the span a-b.  Raises KeyError if none exist.
    Accepted names:
      • "Straight Distance of a-b"
      • "Sag of a-b"
    """
    candidates = [
        f"Straight Distance of {a}-{b}",
        f"Sag of {a}-{b}",
    ]
    for col in candidates:
        if col in df.columns:
            return col
    raise KeyError(f"No distance/sag column found for span {a}-{b}. "
                   f"Tried: {', '.join(candidates)}")


RESULT_SHEET_NAMES = [
    "Result_Dist_Ph-Ph (10ºC)",

]



 
def _which_circuit(tower: int) -> int:
    """
    Return
      1  if tower in CIRCUIT_1
      2  if tower in CIRCUIT_2
      0  otherwise  (earth-wire / unknown circuit)
    """
    if tower in CIRCUIT_1:
        return 1
    if tower in CIRCUIT_2:
        return 2
    return EARTH_WIRE


def _parse_sheet_name(sheet_name: str) -> Tuple[int, int, int, int]:
    """
    Extract digit1, digit2, digit3, digit4 from a name like 41-42_43-44.
    """
    m = SHEET_RE.match(sheet_name)
    if not m:
        raise ValueError(f"Sheet name {sheet_name!r} does not match d1-d2_d3-d4.")
    d1, d2, d3, d4 = map(int, m.groups())
    return d1, d2, d3, d4


def _load_metadata(xl_path: str) -> Tuple[str, str]:
    """
    Extract start & end span codes from the workbook filename since in this case we only have the span names
    on the workbook name.
    Returns (start_span, end_span_suffix).
    """
    m = FILE_RE.search(os.path.basename(xl_path))
    if not m:
        raise ValueError(f"Filename {xl_path!r} does not match expected pattern.")
    return m.group(1), m.group(2)


def process_sheet(df: pd.DataFrame,
                  d1: int, d2: int, d3: int, d4: int) -> pd.DataFrame:
    """
    Add/compute the required columns for a single sheet and return the new DF.
    """
    # ------------------------------------------------------------------
    # 2-3  sag  (max of two distance/sag columns)
    # ------------------------------------------------------------------
    col_sd12 = _pick_distance_col(df, d1, d2)
    col_sd34 = _pick_distance_col(df, d3, d4)

    df["sag"] = df[[col_sd12, col_sd34]].max(axis=1)

    # ------------------------------------------------------------------
    # 2-4  K  =  -(0.1/90)*Beta + 0.75  ->change this linear equation if needed
    # ------------------------------------------------------------------
    df["K"] = -(0.1 / 90.0) * df["Beta Angle (°)"] + 0.75

    # ------------------------------------------------------------------
    # 2-5  constant LK -> length of insulators, here is just an arbitrary number for the sake of calculations
    # ------------------------------------------------------------------
    df["LK"] = 2.0  # metres

    # ------------------------------------------------------------------
    # 2-6  C1 / C2  by circuit comparison of d1 & d3
    # ------------------------------------------------------------------
    cir1 = _which_circuit(d1)
    cir3 = _which_circuit(d3)

    # Earth-wire rule: if either tower is EW (0) we use C1 = 1.93 m only
    if EARTH_WIRE in (cir1, cir3):
        df["C1"] = 1.93
        df["C2"] = 0.0
        # required distance = C1 only (rule supplied by user)
        df["required-distance"] = df["C1"]
    else:
        # both towers belong to real circuits -> old logic
        same_circuit = cir1 == cir3
        if same_circuit:
            df["C1"] = 1.93
            df["C2"] = 0.0
        else:
            df["C1"] = 0.0
            df["C2"] = 2.29
            
        # Make sure every column you use in the formula is numeric just for the sake of calculations
        cols_to_float = ["K", "sag", "LK", "C1", "C2"]
        df[cols_to_float] = df[cols_to_float].apply(
            pd.to_numeric, errors="coerce"
        )

        df["required-distance"] = (
        df["K"] * np.sqrt(df["sag"] + df["LK"])   # k * sqrt(f + LK)
        + df["C1"] + df["C2"]                     # + C1 or C2
        )
    # ------------------------------------------------------------------
    # 2-8  diff = actual - required
    # ------------------------------------------------------------------
    df["diff"] = df["Distance Between Powerlines"] - df["required-distance"]

    # ------------------------------------------------------------------
    # 2-9  worst-case  (single value, other cells left blank)
    # ------------------------------------------------------------------
    worst_idx = df["diff"].idxmin()      # row where diff is the lowest
    worst_val = df.loc[worst_idx, "diff"]

    df["worst-case"] = np.nan            # start with empty cells
    df.loc[worst_idx, "worst-case"] = worst_val
    
    
    
    # ---------------- summary for results workbook -----------------
    # Row with non-nan worst-case value
    worst_idx = df["worst-case"].first_valid_index()
    summary = {
        "row_num": worst_idx + 2,                       # +2 because Excel rows start at 1 and DF header was row 1
        "station": df.loc[worst_idx, "Station"],
        "sag": df.loc[worst_idx, "sag"],
        "lk": df.loc[worst_idx, "LK"],
        "beta": df.loc[worst_idx, "Beta Angle (°)"],
        "k_factor": df.loc[worst_idx, "K"],
        "c1": df.loc[worst_idx, "C1"],
        "c2": df.loc[worst_idx, "C2"],
        "req": df.loc[worst_idx, "required-distance"],
        "cur": df.loc[worst_idx, "Distance Between Powerlines"],
    }
    # diff_req_cur = summary["req"] - summary["cur"]
    # build the flag
    flag = np.where(summary["req"] > summary["cur"], "not ok", "ok")

    # unwrap to ordinary Python str values, then assign
    summary["ok_flag"] = flag.tolist()          # ← crucial step

    return df, summary



# ------------------------------------------------------------------
# helper: open existing results workbook and verify tabs
# ------------------------------------------------------------------
def _open_results_wb(path: str):
    """
    Return (wb, sheets_dict).  Sheets_dict maps the *required* template
    names to openpyxl Worksheet objects.  If a required sheet is missing
    we raise immediately instead of silently creating a new one.
    """
    if not os.path.exists(path):
        raise FileNotFoundError(
            f"Results workbook '{path}' does not exist. "
            "Create it manually with the template sheets first."
        )

    wb = load_workbook(path)

    missing = [name for name in (PH_PH_SHEET, PH_EW_SHEET)
               if name not in wb.sheetnames]
    if missing:
        raise KeyError(
            f"Results workbook is missing sheet(s): {', '.join(missing)}. "
            "Please add them and run again."
        )

    return wb, {name: wb[name] for name in (PH_PH_SHEET, PH_EW_SHEET)}  


# ------------------------------------------------------------------
# MAIN: process a single Cond_10C workbook
# ------------------------------------------------------------------
def process_workbook(xl_path: str) -> None:
    """
    • Adds calculation columns to every matching sheet in *xl_path*.
    • Appends one worst-case summary row to either
        Result_Dist_Ph-Ph (10ºC)   – for phase/phase spans
        Result_Dist_Ph-EW          – for phase / earth-wire spans
      in the existing results workbook RESULTS_PATH.
    • Saves <original>_processed.xlsx next to the source file.
    """

    start_span, end_suffix = _load_metadata(xl_path)
    end_span = f"{start_span[:-3]}{end_suffix}"
    print(f"  Processing {os.path.basename(xl_path)} "
          f"[{start_span} → {end_span}]")

    # ------------ open results workbook (must already exist) ----------
    res_wb, res_sheets = _open_results_wb(RESULTS_PATH)
    next_row = {name: ws.max_row + 1 for name, ws in res_sheets.items()}

    # ------------ open source workbook --------------------------------
    wb = load_workbook(xl_path)

    for sheet_name in wb.sheetnames:
        try:
            d1, d2, d3, d4 = _parse_sheet_name(sheet_name)
        except ValueError:
            continue                                       # skip others

        cir1 = _which_circuit(d1)
        cir3 = _which_circuit(d3)
        ew_span = (0 in (cir1, cir3))                     # earth-wire?

        tag = "Ph-EW" if ew_span else "Ph-Ph"
        print(f"    • {sheet_name}  ({tag})")

        ws = wb[sheet_name]
        df = pd.DataFrame(ws.values)
        df.columns = df.iloc[0]
        df = df.iloc[1:].reset_index(drop=True)

        try:
            df, summ = process_sheet(df, d1, d2, d3, d4)
        except Exception as exc:
            print(f"      !! skipped – {exc}")
            continue                       # keep going with other sheets

        # ---------- choose target sheet and column map -----------------
        if ew_span:
            tgt_ws = res_sheets[PH_EW_SHEET]
            COL    = {
                "row": 1, "src_start": 2, "src_end": 3, "src_set": 4,
                "src_phase": 5, "tgt_start": 6, "tgt_end": 7,
                "tgt_set": 8, "tgt_phase": 9, "station": 10,
                "wind": 11,               # left blank
                "c1": 12, "c3": 13,
                "req": 14, "cur": 15, "ok": 16,
            }
        else:
            tgt_ws = res_sheets[PH_PH_SHEET]
            COL    = {
                "row": 1, "src_start": 2, "src_end": 3, "src_set": 4,
                "src_phase": 5, "tgt_start": 6, "tgt_end": 7,
                "tgt_set": 8, "tgt_phase": 9, "station": 10,
                "sag": 11, "lk": 12, "weather": 13, "beta": 14,
                "w": 15, "k": 16, "c1": 17, "c2": 18,
                "req": 19, "cur": 20, "ok": 21,
            }

        r = next_row[tgt_ws.title]
        next_row[tgt_ws.title] += 1

        def put(key, value=""):
            col = COL.get(key)
            if col:
                tgt_ws.cell(r, col, value)

        # common fields
        put("row",        summ["row_num"])
        put("src_start",  start_span)
        put("src_end",    end_span)
        put("src_set",    d1)
        put("src_phase",  d2)
        put("tgt_start",  start_span)
        put("tgt_end",    end_span)
        put("tgt_set",    d3)
        put("tgt_phase",  d4)
        put("station",    summ["station"])
        put("req",        summ["req"])
        put("cur",        summ["cur"])
        put("ok",         summ["ok_flag"])
        put("c1",         summ["c1"])

        if ew_span:
            put("c3", summ["c2"])
        else:
            put("sag",     summ["sag"])
            put("lk",      summ["lk"])
            put("beta",    summ["beta"])
            put("k",       summ["k_factor"])
            put("c2",      summ["c2"])
            put("weather", 10)
            put("w", "")

        # ------------ overwrite processed data back to sheet ----------
        ws.delete_rows(1, ws.max_row)
        for row in dataframe_to_rows(df, index=False, header=True):
            ws.append(row)

    # ------------ save both workbooks ---------------------------------
    proc_path = os.path.splitext(xl_path)[0] + "_processed.xlsx"
    wb.save(proc_path)
    res_wb.save(RESULTS_PATH)
    print(f"    → written {os.path.basename(proc_path)}")


# ----------------------------------------------------------------------
# ----------------------------------------------------------------------
def main() -> None:
    ap = argparse.ArgumentParser(
        description="Populate sag / clearance data in Cond_10C workbooks.")
    ap.add_argument("input_folder", help="Folder containing the XLSX source files")
    args = ap.parse_args()

    pattern = os.path.join(args.input_folder, "Cond_10C_*.xlsx")
    files = glob.glob(pattern)
    if not files:
        print("No matching files found.")
        return

    for fp in files:
        try:
            process_workbook(fp)
        except Exception as exc:
            print(f"ERROR in {fp}: {exc}")


if __name__ == "__main__":
    main()
