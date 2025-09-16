from pathlib import Path
import re
import sys
from collections import defaultdict

import pandas as pd
from openpyxl import load_workbook, Workbook

# ----------------------------------------------------------------------
# CONFIGURABLE
# ----------------------------------------------------------------------
FOLDER              = Path(r"C:\Users\bimax\DC\ACCDocs\Axpo Grid AG\DEMO_AXPO_Leitungen\Project Files\Grid 4.0 - PLS Distances Development\elham-outputs")   
OUTPUT_FILE         = FOLDER / "Formulas Distances_rev1.xlsx"
EW_PH_REPL          = ["41-21", "42-22", "43-23", "44-24", "45-25", "46-26"]
C_GROUP_1           = {"41", "42", "43"}
C_GROUP_2           = {"44", "45", "46"}

# ----------------------------------------------------------------------
# 1. Locate WindMaster workbooks and companion *Left_* filenames
# ----------------------------------------------------------------------
wm_left  = None
wm_right = None
other_left_files = []                  # EW650-Ph390 , EW390-Ph650 , Ph650-Ph390_XX-YY

for f in FOLDER.glob("*.xlsx"):
    name = f.name
    if re.search(r"WindMasterLeft",  name, re.I):
        wm_left  = f
    elif re.search(r"WindMasterRight", name, re.I):
        wm_right = f
    elif "_Left_EW" in name or "_Left_Ph" in name:
        other_left_files.append(f)

if not (wm_left and wm_right):
    sys.exit("WindMasterLeft/Right files not found.")

# ----------------------------------------------------------------------
# 2. Build sheet-lists
# ----------------------------------------------------------------------
# ----------  NEW  --------------------------------------------------
def classify_family(sheet_name: str) -> str:
    """
    Returns 'EW-Ph' if either d1 or d3 equals 59 or 39, otherwise 'Ph-Ph'.
    Sheet name is expected to contain the two-digit d1 and d3 codes,
    e.g. 41-21_W650_59-39_W390.
    """
    m = re.search(r"(\d{2})-\d{2}_W\d+_(\d{2})-\d{2}", sheet_name)
    if not m:
        return "Ph-Ph"            # fallback – should not happen
    d1, d3 = m.groups()
    return "EW-Ph" if d1 in ("59", "39") or d3 in ("59", "39") else "Ph-Ph"



def swap_sheet_name(sheet_name):
    # Swaps the two wire sections in the sheet name
    m = re.match(r'(.+?_W\d+)_([0-9\-]+_W\d+)', sheet_name)
    if m:
        return f"{m.group(2)}_{m.group(1)}"
    else:
        # If it doesn't match, return original
        return sheet_name


def extract_ew_ph_pressures(sheet_name):
    # Example: "59-39_W390_41-21_W650" -> 390, 650
    m = re.search(r'_W(\d+).*_W(\d+)', sheet_name)
    if m:
        return int(m.group(1)), int(m.group(2))
    return None, None

def extract_ph_ph_pressures(sheet_name):
    # Example: "42-22_W650_41-21_W390" -> 650, 390
    m = re.findall(r'_W(\d+)', sheet_name)
    if m and len(m) >= 2:
        return int(m[0]), int(m[1])
    return None, None

def ew_ph_variants(tag: str) -> list[str]:
    m = re.match(r"EW(\d+)-Ph(\d+)", tag, re.I)
    if not m:
        return []
    ew, ph = m.groups()
    base_ew = f"59-39_W{ew}"
    base_ph = f"W{ph}"
    return [f"{base_ew}_{rep}_{base_ph}" for rep in EW_PH_REPL]

def ph_ph_variants(fname: str) -> list[str]:
    m = re.search(r"Ph(\d+)-Ph(\d+)_(\d{2})-(\d{2})", fname, re.I)
    if not m: return []
    ph1, ph2, d1, d3 = m.groups()
    dcode = f"{d1}-{d3}"
    return [f"{rep}_W{ph1}_{dcode}_W{ph2}" for rep in EW_PH_REPL]

sheet2family = {}          # sheet_name -> "EW-Ph" | "Ph-Ph"
sheet2struct = {}  # sheet_name -> (start, end)

for f in other_left_files:
    # Extract the TR part: TRxxxxxxnnn_mmm or similar
    m = re.search(r"(TR\d+[a-z]?\d{3})_(\d{3})", f.name, re.I)
    if m:
        base = m.group(1)  # e.g. TR1730a001
        end_num = m.group(2)  # e.g. 002
        # build start and end structure strings
        start_structure = base
        # Replace last three digits with end_num
        end_structure = base[:-3] + end_num
    else:
        start_structure = end_structure = ""

    if "_Left_EW" in f.name:
        tag = re.search(r"EW\d+-Ph\d+", f.name)[0]
        for s in ew_ph_variants(tag):
            sheet2family[s] = "EW-Ph"
            sheet2struct[s] = (start_structure, end_structure)

        # --------  NEW – also register the swapped name -------------
        sw = swap_sheet_name(s)
        sheet2family[sw] = "EW-Ph"
        sheet2struct[sw] = (start_structure, end_structure)

    elif "_Left_Ph" in f.name:
        for s in ph_ph_variants(f.name):
            sheet2family[s] = "Ph-Ph"
            sheet2struct[s] = (start_structure, end_structure)

            # --------  NEW ---------------------------------------------
            sw = swap_sheet_name(s)
            sheet2family[sw] = "Ph-Ph"
            sheet2struct[sw] = (start_structure, end_structure)


all_sheets_needed = list(sheet2family.keys())
print("Total sheets that should be processed:", len(all_sheets_needed))
for s in all_sheets_needed:
    print("   •", s)

# ----------------------------------------------------------------------
# 3. Make sure sheets exist in both workbooks
# ----------------------------------------------------------------------
def ensure_sheet(book: load_workbook, sheet_name: str):
    if sheet_name not in book.sheetnames:
        ws = book.create_sheet(title=sheet_name)
        ws.append(["Station", "Sag of ???", "Distance Between Powerlines"])

left_wb  = load_workbook(wm_left)
right_wb = load_workbook(wm_right)

for s in all_sheets_needed:
    ensure_sheet(left_wb,  s)
    ensure_sheet(right_wb, s)

left_wb.save(wm_left)
right_wb.save(wm_right)

# ----------------------------------------------------------------------
# 3.1 Helper to compute C3, C4, Required distance
# ----------------------------------------------------------------------
def add_distance_columns(df: pd.DataFrame, sheet_name: str) -> pd.DataFrame:
    m = re.search(r"(\d{2})-\d{2}.*_W\d+_(\d{2})-\d{2}", sheet_name)
    if m:
        d1, d3 = m.groups()
    else:
        m2 = re.match(r"59-39_W\d+_(\d{2})-\d{2}_W\d+", sheet_name)
        if not m2:
            raise ValueError(f"Cannot parse digits from {sheet_name}")
        d1, d3 = m2.group(1), m2.group(1)

    same_circuit = (
        (d1 in C_GROUP_1 and d3 in C_GROUP_1) or
        (d1 in C_GROUP_2 and d3 in C_GROUP_2)
    )
    df["C3"] = 1.35 if same_circuit else 0
    df["C4"] = 0 if same_circuit else 1.60
    df["Required distance"] = df[["C3", "C4"]].max(axis=1)
    return df

print("Reading both WindMaster workbooks in one go …")
left_dfs  = pd.read_excel(wm_left,  sheet_name=None, engine="openpyxl")   # dict
right_dfs = pd.read_excel(wm_right, sheet_name=None, engine="openpyxl")

winners = []                                    # results accumulator
total   = len(all_sheets_needed)
for idx, sheet in enumerate(all_sheets_needed, 1):
    print(f"[{idx:>3}/{total}] processing sheet «{sheet}»")
    m_digits = re.search(r"(\d{2})-\d{2}_W\d+_(\d{2})-\d{2}", sheet)
    if m_digits and m_digits.group(1) == m_digits.group(2):
        print("   · symmetrical sheet – skipped")
        continue

    # Generate both sheet name variants
    variant_a = sheet
    variant_b = swap_sheet_name(sheet)
    sheet_variants = [variant_a, variant_b]

    entry = defaultdict(lambda: None)
    min_distances = []
    min_rows = []
    min_sources = []

    for side, dfs in (("Left", left_dfs), ("Right", right_dfs)):
        for variant in sheet_variants:
            try:
                df = dfs.get(variant)
                if df is None or df.empty:
                    continue

                if "Required distance" not in df.columns:
                    df = add_distance_columns(df.copy(), variant)
                    dfs[variant] = df

                if df["Required distance"].dropna().empty:
                    continue

                # Get minimum Distance Between Powerlines from the whole column
                dist_col = df["Distance Between Powerlines"].dropna()
                if dist_col.empty:
                    continue

                min_dist = dist_col.min()
                min_distances.append((min_dist, side, variant))

                # Get index of row with minimum Distance Between Powerlines
                min_idx = dist_col.idxmin()
                min_row = df.loc[min_idx]
                min_rows.append(min_row)
                min_sources.append((side, variant))

            except Exception as exc:
                print(f"   ⚠️  {side}::{variant} skipped – {exc}")
                continue

    # Select the winner: the sheet/side with the lowest minimum distance
    if not min_distances:
        print("   · sheet contains no data on either side – skipped")
        continue

    winner_tuple = min(min_distances, key=lambda x: x[0])  # (min_dist, side, variant)
    winner_dist, winner_side, winner_variant = winner_tuple

    # Find the corresponding row for the winner
    winner_row = None
    for i, (side, variant) in enumerate(min_sources):
        if side == winner_side and variant == winner_variant:
            winner_row = min_rows[i]
            break

    if winner_row is None:
        print("   · winner row not found – skipped")
        continue

    # Fill entry dict with winner info
    entry.update({
        "winner"        : winner_side,
        "sheet"         : winner_variant,
        "family"        : sheet2family.get(winner_variant, ""),
        "station"       : winner_row["Station"],
        "c3"            : winner_row.get("C3", None),
        "c4"            : winner_row.get("C4", None),
        "current_dist"  : winner_dist,
        "ok"            : "OK" if winner_dist >= (winner_row.get("C3", 0) if winner_row.get("C3", 0) else winner_row.get("C4", 0)) else "NO OK",
    })

    # For wind pressure columns
    entry["wind_pressure"] = int(re.search(r"_W(\d+)", winner_variant).group(1)) if re.search(r"_W(\d+)", winner_variant) else None

    winners.append(entry)
    print("Sheets examined :", len(all_sheets_needed))
    print("Winners found   :", len(winners))




# ------------------------------------------------------------------
# 4b.  Write modified DataFrames back to disk ONCE per workbook
# ------------------------------------------------------------------
def flush_back(xlsx_path: Path, dfs_dict: dict):
    with pd.ExcelWriter(xlsx_path, engine="openpyxl", mode="w") as w:
        for sh, df in dfs_dict.items():
            df.to_excel(w, sheet_name=sh, index=False)

print("Writing updated sheets back to the two workbooks …")
flush_back(wm_left,  left_dfs)
flush_back(wm_right, right_dfs)
print("Sheet processing finished.\n")

# ----------------------------------------------------------------------
# 5. Append results to output workbook (ALWAYS create new -Auto sheets)
# ----------------------------------------------------------------------
FIXED_HEADERS_EW_PH = [
    "Source Start Structure",
    "Source End Structure",
    "Target Start Structure",
    "Target End Structure",
    "Source Set", 
    "Target Set",
    "Station (m)",
    "Earth wire wind pressure",
    "Phase wire wind pressure",
    "C3",
    "Required Distance At Min. (m)",
    "Current distance (m)", 
    "OK / NO OK",
    "Source Sheet"
]

FIXED_HEADERS_PH_PH = [
    "Source Start Structure",
    "Source End Structure",
    "Target Start Structure",
    "Target End Structure",
    "Source Set", 
    "Target Set",
    "Station (m)",
    "first wire wind pressure",
    "second wire wind pressure",
    "C3",
    "C4",
    "Required Distance At Min. (m)",
    "Current distance (m)", 
    "OK / NO OK",
    "Source Sheet"
]


def append_rows(book_path: Path, sheet_name: str, rows: list[list], headers: list[str]):
    if book_path.exists():
        wb = load_workbook(book_path)
    else:
        wb = Workbook()
        wb.remove(wb.active)

    base = f"{sheet_name}-Auto"
    candidate = base
    n = 1
    while candidate in wb.sheetnames:
        candidate = f"{base}({n})"
        n += 1
    ws = wb.create_sheet(candidate)

    for c, hdr in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=hdr)

    for r_idx, data in enumerate(rows, 2):
        for c_idx, value in enumerate(data, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    wb.save(book_path)
    print(f"   → wrote {len(rows)} row(s) to «{candidate}» in {book_path.name}")


# ----------------------------------------------------------------------
# build the output rows -------------------------------------------------
rows_EW_Ph, rows_Ph_Ph = [], []

for w in winners:
    # --------------------------------------------------------------
    # decide the family *again* with the new rule (overrides mapping)
    # --------------------------------------------------------------
    family = classify_family(w["sheet"])

    m_digits = re.search(r"(\d{2})-\d{2}_W\d+_(\d{2})-\d{2}", w["sheet"])
    d1, d3 = (m_digits.groups() if m_digits else ("", ""))

    # look up structures – guaranteed to exist after patch #2
    sheet_name = w["sheet"]                       # the winner sheet
    start_struct, end_struct = sheet2struct.get(sheet_name, (None, None))

    if start_struct is None:                      # not found → try swap
        sw_name = swap_sheet_name(sheet_name)
        start_struct, end_struct = sheet2struct.get(sw_name, ("", ""))


    if family == "EW-Ph":
        # --- EW-Ph ------------------------------------------------------
        ew_wp, ph_wp = extract_ew_ph_pressures(sheet_name)
        required_distance = 1.35                    # always the same

        row = [
            start_struct, end_struct,               # Source
            start_struct, end_struct,               # Target
            d1, d3,
            w["station"],
            ew_wp,
            ph_wp,
            1.35,                                   # C3 column
            required_distance,                      # Required Distance col
            w["current_dist"],
            w["ok"],
            sheet_name,
        ]
        rows_EW_Ph.append(row)

    else:
        # --- Ph-Ph ------------------------------------------------------
        fw_wp, sw_wp = extract_ph_ph_pressures(sheet_name)

        # choose the non-zero coefficient
        required_distance = w["c3"] if w["c3"] and w["c3"] != 0 else w["c4"]

        row = [
            start_struct, end_struct,
            start_struct, end_struct,
            d1, d3,
            w["station"],
            fw_wp,
            sw_wp,
            w["c3"] or "",                          # C3 column
            w["c4"] or "",                          # C4 column
            required_distance,                      # Required Distance col
            w["current_dist"],
            w["ok"],
            sheet_name,
        ]
        rows_Ph_Ph.append(row)



# ----------------------------------------------------------------------
# write to workbook (ALWAYS create new -Auto sheets) -------------------
print("EW-Ph rows :", len(rows_EW_Ph))
print("Ph-Ph rows :", len(rows_Ph_Ph))
print("Sheets examined :", len(all_sheets_needed))
print("Winners found   :", len(winners))


if rows_EW_Ph:
    append_rows(OUTPUT_FILE, "Result_Dist_EW-Ph(Wind)", rows_EW_Ph, FIXED_HEADERS_EW_PH)

if rows_Ph_Ph:
    append_rows(OUTPUT_FILE, "Result_Dist_Ph-Ph(Wind)", rows_Ph_Ph, FIXED_HEADERS_PH_PH)



print("Done. Results appended to", OUTPUT_FILE.resolve())
